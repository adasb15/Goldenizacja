"""Logika biznesowa warstwy."""
import csv
import hashlib
import io
import json
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

try:
    from sqlalchemy.exc import IntegrityError
except ImportError:
    class IntegrityError(Exception):
        pass

from app.layers.ingestion.schemas import RawLoadResponse, RelationalQueryInfo


SUPPORTED_FILE_TYPES = {"csv", "json", "xml", "xlsx"}
SUPPORTED_SOURCE_SYSTEMS = {
    "CEIDG": "Centralna Ewidencja i Informacja o Dzialalnosci Gospodarczej",
    "KRS": "Krajowy Rejestr Sadowy",
    "REGON": "Rejestr REGON",
    "VAT": "Wykaz podatnikow VAT",
    "PESEL": "Rejestr PESEL",
    "KNF_AGENT": "KNF Rejestr posrednikow ubezpieczeniowych - agent",
    "KNF_PRACOWNIK_AGENTA": "KNF Rejestr posrednikow ubezpieczeniowych - pracownik agenta",
    "KNF_FIRMY_INWESTYCYJNE": "KNF Rejestr firm inwestycyjnych",
    "KNF_PIENIADZ_ELEKTRONICZNY": "KNF Rejestr dostawcow i wydawcow pieniadza elektronicznego",
    "GLEIF": "GLEIF",
    "INSURANCE_CORE": "Oracle Insurance Core - relacyjne zrodlo przez ODBC",
}

INSURANCE_CORE_PARTY_MAIN_SQL = """
SELECT
    c.CUST_UID,
    c.EXT_REF_NO,
    c.SUBJECT_KIND,
    c.PARTY_LABEL,
    c.BRAND_LABEL,
    c.FORM_CD,
    c.LIFE_CYCLE_CD,
    c.ACTIVATION_DT,
    c.TERMINATION_DT,
    c.RISK_BUCKET,
    c.SEGMENT_HINT,
    c.LEGACY_RANK,
    c.SCREENING_FLAG,
    c.SOURCE_REVISION,
    c.NOTE_TXT,
    c.LAST_TOUCH_TS,
    b.HOME_TOWN AS UNIT_TOWN_HINT,
    b.AREA_BUCKET AS AREA_BUCKET,
    b.OPS_NOTE AS UNIT_OPS_NOTE,
    nip.REF_VALUE AS TAX_REF,
    regon.REF_VALUE AS STAT_REG_REF,
    krs.REF_VALUE AS COURT_REF,
    addr.LINE_1 AS ADDR_TXT,
    addr.LINE_2 AS ADDR_EXTRA_TXT,
    addr.CITY_HINT AS MUNICIPAL_UNIT,
    addr.POST_AREA AS POST_AREA,
    addr.ISO_MARKET AS ISO_MARKET,
    addr.GEO_ZONE AS GEO_ZONE,
    addr.COURIER_ROUTE AS COURIER_ROUTE,
    email.CONTACT_PAYLOAD AS MAILBOX,
    phone.CONTACT_PAYLOAD AS TEL_NOTE,
    iban.CLEARING_REF AS SETTLEMENT_ACC,
    contracts.ACTIVE_CONTRACTS_COUNT AS POLICY_ACTIVE_QTY,
    contracts.LATEST_CONTRACT_STATUS AS POLICY_LAST_STATE
FROM CLIENT_ACCOUNT c
LEFT JOIN ORG_UNIT b
    ON b.ORG_UNIT_ID = c.ORG_UNIT_ID
LEFT JOIN CUSTOMER_IDENTIFIER nip
    ON nip.CUST_UID = c.CUST_UID
   AND nip.REF_KIND = 'NIP'
LEFT JOIN CUSTOMER_IDENTIFIER regon
    ON regon.CUST_UID = c.CUST_UID
   AND regon.REF_KIND = 'REGON'
LEFT JOIN CUSTOMER_IDENTIFIER krs
    ON krs.CUST_UID = c.CUST_UID
   AND krs.REF_KIND = 'KRS'
LEFT JOIN CUSTOMER_ADDRESS addr
    ON addr.CUST_UID = c.CUST_UID
   AND addr.LOC_ROLE = 'REGISTERED'
LEFT JOIN CLIENT_CONTACT_LOG email
    ON email.CUST_UID = c.CUST_UID
   AND email.CHANNEL_CODE = 'EMAIL'
   AND email.IS_CONFIRMED = 1
LEFT JOIN CLIENT_CONTACT_LOG phone
    ON phone.CUST_UID = c.CUST_UID
   AND phone.CHANNEL_CODE = 'PHONE'
   AND phone.IS_CONFIRMED = 1
LEFT JOIN PAYMENT_ACCOUNT iban
    ON iban.CUST_UID = c.CUST_UID
   AND iban.IS_ACTIVE = 1
LEFT JOIN (
    SELECT
        cp.CUST_UID,
        COUNT(*) AS ACTIVE_CONTRACTS_COUNT,
        MAX(co.DEAL_STATE) AS LATEST_CONTRACT_STATUS
    FROM CONTRACT_PARTY cp
    JOIN CONTRACT co
        ON co.DEAL_ID = cp.DEAL_ID
    WHERE co.DEAL_STATE = 'ACTIVE'
    GROUP BY cp.CUST_UID
) contracts
    ON contracts.CUST_UID = c.CUST_UID
WHERE c.SUBJECT_KIND = 'ORG'
"""

INSURANCE_CORE_PERSON_ROLES_SQL = """
SELECT
    ar.CUST_UID,
    ar.GIVEN_TXT,
    ar.FAMILY_TXT,
    ar.NATIONAL_REF,
    ar.LICENSE_TOKEN,
    ar.ROLE_BUCKET,
    ar.VALID_SINCE,
    ar.VALID_UNTIL,
    ar.HR_SOURCE,
    ar.QUALITY_NOTE
FROM AGENT_ASSIGNMENT ar
"""

INSURANCE_CORE_RELATED_PARTIES_SQL = """
SELECT
    rc.PARENT_UID AS CUST_UID,
    child.EXT_REF_NO AS RELATED_EXT_REF_NO,
    child.PARTY_LABEL AS RELATED_PARTY_LABEL,
    rc.LINK_BUCKET,
    rc.VALID_SINCE,
    rc.VALID_UNTIL,
    rc.ORIGIN_HINT,
    rc.COMMENT_TXT
FROM RELATED_CLIENT rc
JOIN CLIENT_ACCOUNT child
    ON child.CUST_UID = rc.CHILD_UID
"""

INSURANCE_CORE_PERSON_MAIN_SQL = """
SELECT
    'AGENT-' || ar.ASSIGNMENT_ID AS PERSON_REF,
    ar.NATIONAL_REF,
    ar.GIVEN_TXT,
    ar.FAMILY_TXT,
    ar.LICENSE_TOKEN,
    ar.ROLE_BUCKET,
    ar.VALID_SINCE,
    ar.VALID_UNTIL,
    ar.HR_SOURCE,
    ar.QUALITY_NOTE,
    parent.EXT_REF_NO AS RELATED_PARTY_REF,
    parent.PARTY_LABEL AS RELATED_PARTY_LABEL,
    NULL AS SECOND_GIVEN_TXT,
    NULL AS BIRTH_DT_HINT,
    NULL AS BIRTH_PLACE_HINT,
    NULL AS GENDER_HINT,
    NULL AS CITIZENSHIP_HINT,
    NULL AS MAILBOX,
    NULL AS TEL_NOTE,
    NULL AS ADDR_TXT,
    NULL AS MUNICIPAL_UNIT,
    NULL AS POST_AREA,
    NULL AS ISO_MARKET
FROM AGENT_ASSIGNMENT ar
JOIN CLIENT_ACCOUNT parent
    ON parent.CUST_UID = ar.CUST_UID
UNION ALL
SELECT
    c.EXT_REF_NO AS PERSON_REF,
    NULL AS NATIONAL_REF,
    SUBSTR(c.PARTY_LABEL, 1, INSTR(c.PARTY_LABEL || ' ', ' ') - 1) AS GIVEN_TXT,
    SUBSTR(c.PARTY_LABEL, INSTR(c.PARTY_LABEL || ' ', ' ') + 1) AS FAMILY_TXT,
    NULL AS LICENSE_TOKEN,
    c.SEGMENT_HINT AS ROLE_BUCKET,
    c.ACTIVATION_DT AS VALID_SINCE,
    c.TERMINATION_DT AS VALID_UNTIL,
    c.SOURCE_REVISION AS HR_SOURCE,
    c.NOTE_TXT AS QUALITY_NOTE,
    NULL AS RELATED_PARTY_REF,
    NULL AS RELATED_PARTY_LABEL,
    NULL AS SECOND_GIVEN_TXT,
    NULL AS BIRTH_DT_HINT,
    NULL AS BIRTH_PLACE_HINT,
    NULL AS GENDER_HINT,
    NULL AS CITIZENSHIP_HINT,
    email.CONTACT_PAYLOAD AS MAILBOX,
    phone.CONTACT_PAYLOAD AS TEL_NOTE,
    addr.LINE_1 AS ADDR_TXT,
    addr.CITY_HINT AS MUNICIPAL_UNIT,
    addr.POST_AREA AS POST_AREA,
    addr.ISO_MARKET AS ISO_MARKET
FROM CLIENT_ACCOUNT c
LEFT JOIN CLIENT_CONTACT_LOG email
    ON email.CUST_UID = c.CUST_UID
   AND email.CHANNEL_CODE = 'EMAIL'
   AND email.IS_CONFIRMED = 1
LEFT JOIN CLIENT_CONTACT_LOG phone
    ON phone.CUST_UID = c.CUST_UID
   AND phone.CHANNEL_CODE = 'PHONE'
   AND phone.IS_CONFIRMED = 1
LEFT JOIN CUSTOMER_ADDRESS addr
    ON addr.CUST_UID = c.CUST_UID
   AND addr.LOC_ROLE = 'REGISTERED'
WHERE c.SUBJECT_KIND = 'PERSON'
"""

GENERIC_INSURANCE_CORE_QUERY_NAME = "insurance_core_export"
INSURANCE_CORE_QUERY_BY_ENTITY_TYPE = {
    "PARTY": "insurance_core_party_export",
    "PERSON": "insurance_core_person_export",
}

RELATIONAL_QUERY_DEFINITIONS = {
    "insurance_core_party_export": {
        "source_system_code": "INSURANCE_CORE",
        "entity_type": "PARTY",
        "description": "Eksport podmiotow z relacyjnego systemu Oracle Insurance Core",
        "file_name": "insurance_core_party_export.json",
        "main_sql": INSURANCE_CORE_PARTY_MAIN_SQL,
        "person_roles_sql": INSURANCE_CORE_PERSON_ROLES_SQL,
        "related_parties_sql": INSURANCE_CORE_RELATED_PARTIES_SQL,
    },
    "insurance_core_person_export": {
        "source_system_code": "INSURANCE_CORE",
        "entity_type": "PERSON",
        "description": "Eksport osob z relacyjnego systemu Oracle Insurance Core",
        "file_name": "insurance_core_person_export.json",
        "main_sql": INSURANCE_CORE_PERSON_MAIN_SQL,
    },
}



class UnsupportedFileTypeError(ValueError):
    pass

class UnsupportedSourceSystemError(ValueError):
    pass


class InvalidFileContentError(ValueError):
    pass


class UnsupportedRelationalQueryError(ValueError):
    pass


class RelationalConnectionConfigurationError(ValueError):
    pass


def _get_file_type(filename: str) -> str:
    # Wyciągamy typ z rozszerzenia, żeby zapisać go w RAW i dobrać późniejszy parser stagingu
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix not in SUPPORTED_FILE_TYPES:
        raise UnsupportedFileTypeError(
            f"Unsupported file type '{suffix}'. Supported types: {', '.join(sorted(SUPPORTED_FILE_TYPES))}."
        )
    return suffix


def _validate_source_system_code(source_system_code: str) -> str:
    # Normalizujemy kod źródła, żeby użytkownik w Postmanie nie musiał pilnować wielkości liter
    normalized = source_system_code.strip().upper()
    if normalized not in SUPPORTED_SOURCE_SYSTEMS:
        raise UnsupportedSourceSystemError(
            f"Unsupported source system '{source_system_code}'."
        )
    return normalized



def _validate_and_count_records(file_type: str, content: bytes) -> int | None:
    if not content:
        raise InvalidFileContentError("Uploaded file is empty.")

    if file_type == "csv":
        # Liczymy rekordy CSV bez nagłówka, żeby raw-load raportował rozmiar paczki przed mapowaniem
        text = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            raise InvalidFileContentError("CSV file has no rows.")
        return max(len(rows) - 1, 0)

    if file_type == "json":
        # Obsługujemy obiekt i listę JSON, żeby raw-load przyjął pojedynczy rekord albo paczkę
        parsed = json.loads(content.decode("utf-8-sig"))
        if isinstance(parsed, list):
            return len(parsed)
        if isinstance(parsed, dict):
            return 1
        raise InvalidFileContentError("JSON root must be an object or an array.")

    if file_type == "xml":
        # Liczymy elementy XML pierwszego poziomu, żeby RAW miał szybki licznik przed dokładnym stagingiem
        root = ElementTree.fromstring(content)
        return len(list(root))

    if file_type == "xlsx":
        try:
            from openpyxl import load_workbook

            # Liczymy niepuste wiersze XLSX, żeby formatowanie arkusza nie zawyżało liczby rekordów
            workbook = load_workbook(
                filename=io.BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise InvalidFileContentError("Invalid XLSX file.") from exc

        sheet = workbook.active
        row_count = 0

        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                row_count += 1

        workbook.close()

        if row_count == 0:
            raise InvalidFileContentError("XLSX file has no rows.")

        return max(row_count - 1, 0)

    return None


def list_relational_queries() -> list[RelationalQueryInfo]:
    return [
        RelationalQueryInfo(
            query_name=GENERIC_INSURANCE_CORE_QUERY_NAME,
            source_system_code="INSURANCE_CORE",
            entity_type="AUTO",
            description="Eksport relacyjny Oracle Insurance Core; entity_type wybiera PARTY, PERSON albo oba w Airflow",
        )
    ]


def import_raw_file(
    db: Any,
    filename: str,
    content: bytes,
    source_system_code: str,
    created_by: str | None = None,
) -> RawLoadResponse:
    # Importujemy plik do RAW bez czyszczenia, żeby zachować oryginalne dane dla kolejnych warstw
    source_system_code = _validate_source_system_code(source_system_code)
    file_type = _get_file_type(filename)
    records_in = _validate_and_count_records(file_type, content)

    return persist_raw_content(
        db=db,
        filename=filename,
        file_type=file_type.upper(),
        content=content,
        source_system_code=source_system_code,
        created_by=created_by,
        records_in=records_in,
    )


def import_relational_source(
    db: Any,
    source_system_code: str,
    query_name: str,
    entity_type: str | None = None,
    created_by: str | None = None,
    connector: Any | None = None,
    repo: Any | None = None,
) -> RawLoadResponse:
    source_system_code = _validate_source_system_code(source_system_code)
    definition = resolve_relational_query_definition(
        source_system_code=source_system_code,
        query_name=query_name,
        entity_type=entity_type,
    )

    records = extract_relational_records(definition, connector=connector)
    content = json.dumps(records, ensure_ascii=False, default=str).encode("utf-8")
    return persist_raw_content(
        db=db,
        filename=str(definition["file_name"]),
        file_type="JSON",
        content=content,
        source_system_code=source_system_code,
        created_by=created_by,
        records_in=len(records),
        repo=repo,
    )


def resolve_relational_query_definition(
    source_system_code: str,
    query_name: str,
    entity_type: str | None = None,
) -> dict[str, Any]:
    normalized_query_name = query_name.strip().casefold()
    normalized_entity_type = entity_type.strip().upper() if entity_type else None

    if normalized_query_name == GENERIC_INSURANCE_CORE_QUERY_NAME:
        if normalized_entity_type not in INSURANCE_CORE_QUERY_BY_ENTITY_TYPE:
            raise UnsupportedRelationalQueryError(
                "Query 'insurance_core_export' requires entity_type PARTY or PERSON."
            )
        normalized_query_name = INSURANCE_CORE_QUERY_BY_ENTITY_TYPE[normalized_entity_type]

    definition = RELATIONAL_QUERY_DEFINITIONS.get(normalized_query_name)
    if definition is None:
        raise UnsupportedRelationalQueryError(f"Unsupported relational query '{query_name}'.")
    if normalized_entity_type and definition["entity_type"] != normalized_entity_type:
        raise UnsupportedRelationalQueryError(
            f"Query '{query_name}' returns '{definition['entity_type']}', not '{normalized_entity_type}'."
        )
    if definition["source_system_code"] != source_system_code:
        raise UnsupportedRelationalQueryError(
            f"Query '{query_name}' is configured for source system "
            f"'{definition['source_system_code']}', not '{source_system_code}'."
        )

    return definition


def extract_relational_records(
    definition: dict[str, Any],
    connector: Any | None = None,
) -> list[dict[str, Any]]:
    connection_config = None
    if connector is None:
        from app.core.config import settings

        connection_config = settings
    if connector is None and connection_config is None:
        raise RelationalConnectionConfigurationError("Oracle connection is not configured.")

    connection = (
        connector(connection_config)
        if connector is not None
        else connect_relational_source(connection_config)
    )
    try:
        main_records = fetch_relational_rows(connection, str(definition["main_sql"]))
        if definition.get("entity_type") != "PARTY":
            return main_records

        person_roles = fetch_relational_rows(connection, str(definition["person_roles_sql"]))
        related_parties = fetch_relational_rows(connection, str(definition["related_parties_sql"]))
    finally:
        close = getattr(connection, "close", None)
        if callable(close):
            close()

    records_by_client_id = {str(record["CUST_UID"]): record for record in main_records}

    roles_by_client_id: dict[str, list[dict[str, Any]]] = {}
    for role in person_roles:
        client_id = str(role.get("CUST_UID"))
        if client_id not in records_by_client_id:
            continue
        roles_by_client_id.setdefault(client_id, []).append(
            {
                "role_group": role.get("ROLE_BUCKET"),
                "first_name": role.get("GIVEN_TXT"),
                "last_name": role.get("FAMILY_TXT"),
                "pesel": role.get("NATIONAL_REF"),
                "license_no": role.get("LICENSE_TOKEN"),
                "valid_from": role.get("VALID_SINCE"),
                "valid_to": role.get("VALID_UNTIL"),
                "source": role.get("HR_SOURCE"),
                "quality_note": role.get("QUALITY_NOTE"),
            }
        )

    parties_by_client_id: dict[str, list[dict[str, Any]]] = {}
    for relation in related_parties:
        client_id = str(relation.get("CUST_UID"))
        if client_id not in records_by_client_id:
            continue
        parties_by_client_id.setdefault(client_id, []).append(
            {
                "relationship_group": relation.get("LINK_BUCKET"),
                "related_source_record_id": relation.get("RELATED_EXT_REF_NO"),
                "name": relation.get("RELATED_PARTY_LABEL"),
                "valid_from": relation.get("VALID_SINCE"),
                "valid_to": relation.get("VALID_UNTIL"),
                "source": relation.get("ORIGIN_HINT"),
                "comment": relation.get("COMMENT_TXT"),
            }
        )

    for client_id, record in records_by_client_id.items():
        if client_id in roles_by_client_id:
            record["RELATED_PERSONS_JSON"] = json.dumps(
                roles_by_client_id[client_id],
                ensure_ascii=False,
                default=str,
            )
        if client_id in parties_by_client_id:
            record["RELATED_PARTIES_JSON"] = json.dumps(
                parties_by_client_id[client_id],
                ensure_ascii=False,
                default=str,
            )

    return list(records_by_client_id.values())


def connect_relational_source(settings: Any) -> Any:
    if settings.oracle_odbc_connection_string:
        return connect_odbc(settings.oracle_odbc_connection_string)

    return connect_oracle_thin(
        user=settings.oracle_app_user,
        password=settings.oracle_app_password,
        host=settings.oracle_host,
        port=settings.oracle_port,
        service_name=settings.oracle_service_name,
    )


def connect_odbc(connection_string: str | None) -> Any:
    if not connection_string:
        raise RelationalConnectionConfigurationError(
            "ORACLE_ODBC_CONNECTION_STRING is not configured."
        )

    import pyodbc

    return pyodbc.connect(connection_string)


def connect_oracle_thin(
    user: str,
    password: str,
    host: str,
    port: int,
    service_name: str,
) -> Any:
    import oracledb

    dsn = f"{host}:{port}/{service_name}"
    return oracledb.connect(user=user, password=password, dsn=dsn)


def fetch_relational_rows(connection: Any, sql: str) -> list[dict[str, Any]]:
    cursor = connection.cursor()
    try:
        cursor.execute(sql)
        columns = [str(column[0]) for column in cursor.description]
        return [
            dict(zip(columns, row))
            for row in cursor.fetchall()
        ]
    finally:
        close = getattr(cursor, "close", None)
        if callable(close):
            close()


def persist_raw_content(
    db: Any,
    filename: str,
    file_type: str,
    content: bytes,
    source_system_code: str,
    created_by: str | None,
    records_in: int | None,
    repo: Any | None = None,
) -> RawLoadResponse:
    if repo is None:
        from app.layers.ingestion.repository import IngestionRepository

    repo = repo or IngestionRepository(db)
    batch = None
    process_log = None

    try:
        source = repo.get_or_create_source_system(
            source_system_code,
            SUPPORTED_SOURCE_SYSTEMS[source_system_code],
        )
        # Tworzymy batch przed zapisem pliku, żeby każdy błąd miał wspólny kontekst importu
        batch = repo.create_import_batch(source.SourceSystem_ID, created_by)
        batch = repo.update_import_batch_status(batch, "PROCESSING")

        process_log = repo.create_process_log(batch.ImportBatch_ID)

        # Liczymy hash pliku, żeby zablokować przypadkowy duplikat identycznego RAW
        file_hash = hashlib.sha256(content).hexdigest()
        raw_file = repo.insert_raw_file(
            import_batch_id=batch.ImportBatch_ID,
            file_name=filename,
            file_type=file_type.upper(),
            file_size=len(content),
            file_hash=file_hash,
            file_content=content,
        )

        repo.finish_process_log(
            process_log,
            status="SUCCESS",
            raw_file_id=raw_file.RawFile_ID,
            records_in=records_in,
            records_out=records_in,
        )
        batch = repo.update_import_batch_status(batch, "RAW_LOADED", finish=True)

        return RawLoadResponse(
            import_batch_id=batch.ImportBatch_ID,
            raw_file_id=raw_file.RawFile_ID,
            file_name=raw_file.File_Name,
            file_type=raw_file.File_Type,
            file_size=raw_file.File_Size,
            file_hash=raw_file.File_Hash,
            records_in=records_in,
            import_status=batch.Import_Status,
        )

    except IntegrityError as exc:
        if hasattr(db, "rollback"):
            db.rollback()
        # Zamieniamy konflikt hasha na błąd walidacji, żeby użytkownik wiedział że plik już istnieje
        error_message = "File with this hash already exists."

        if process_log is not None:
            repo.finish_process_log(process_log, status="FAILED", error_message=error_message)
        if batch is not None:
            repo.update_import_batch_status(batch, "FAILED", error_message=error_message, finish=True)

        raise InvalidFileContentError(error_message) from exc

    except Exception as exc:
        if hasattr(db, "rollback"):
            db.rollback()
        # Domykamy logi przy błędzie technicznym, żeby baza nie zostawiała importu w PROCESSING
        error_message = str(exc)

        if process_log is not None:
            repo.finish_process_log(process_log, status="FAILED", error_message=error_message)
        if batch is not None:
            repo.update_import_batch_status(batch, "FAILED", error_message=error_message, finish=True)

        raise
