import csv
import io
import json
import re
from collections import Counter
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any
from xml.etree import ElementTree

from app.layers.staging_validation.mapper import (
    MISSING_COLUMNS_KEY,
    PARTY_CANONICAL_COLUMNS,
    PERSON_CANONICAL_COLUMNS,
    UNRECOGNIZED_COLUMNS_KEY,
    map_records_to_canonical,
    normalize_entity_type,
)


SOURCE_RECORD_ID_CANDIDATES = (
    "Source_Record_ID",
    "source_record_id",
    "id",
    "ID",
    "firma.id",
    "numerKRS",
    "PESEL",
    "LEI",
)
SOURCE_RECORD_ID_CANDIDATE_KEYS = {
    candidate.strip().casefold()
    for candidate in SOURCE_RECORD_ID_CANDIDATES
}
# Wskazujemy daty PARTY do normalizacji, żeby staging trzymał DATE niezależnie od formatu źródła
PARTY_DATE_COLUMNS = {
    "Establishment_Date",
    "Registration_Date",
    "Deregistration_Date",
    "Decision_Date",
    "Last_Update_Date",
    "Next_Renewal_Date",
    "Direct_Parent_Relationship_Start_Date",
    "Direct_Parent_Relationship_End_Date",
    "Ultimate_Parent_Relationship_Start_Date",
    "Ultimate_Parent_Relationship_End_Date",
}

# Mapujemy prefiksy KRS na role, żeby później zasilić factless person-party role
KRS_PERSON_ROLE_PREFIXES = {
    "CzlonekZarzadu": "BOARD_MEMBER",
    "Prokurent": "PROXY",
    "WspolnikOsoba": "PERSON_SHAREHOLDER",
    "Likwidator": "LIQUIDATOR",
    "CzlonekRadyNadzorczej": "SUPERVISORY_BOARD_MEMBER",
}
KRS_PERSON_ROLE_COLUMN_RE = re.compile(
    r"^(CzlonekZarzadu|Prokurent|WspolnikOsoba|Likwidator|CzlonekRadyNadzorczej)"
    r"(\d+)_(Imie|Nazwisko|PESEL|Funkcja|DataOd|DataDo)$"
)
# Rozpoznajemy wspólników podmiotowych KRS, żeby zapisać ich jako relacje party-party
KRS_PARTY_RELATIONSHIP_COLUMN_RE = re.compile(
    r"^(WspolnikPodmiot)(\d+)_(Nazwa|KRS|NIP|DataOd|DataDo)$"
)
ADDRESS_PREFIX_RE = re.compile(r"^(?:ul\.?|ulica|al\.?|aleja)\s+", re.IGNORECASE)
APARTMENT_PREFIX_RE = re.compile(r"\s+(?:m\.?|lok\.?|lokal)\s+", re.IGNORECASE)
FULL_ADDRESS_STREET_FIRST_RE = re.compile(
    r"^\s*(?P<street_part>.+?),\s*(?P<postal_code>\d{2}-\d{3})\s+(?P<city>.+?)\s*$"
)
FULL_ADDRESS_POSTAL_FIRST_RE = re.compile(
    r"^\s*(?P<postal_code>\d{2}-\d{3})\s+(?P<city>[^,]+?),\s*(?P<street_part>.+?)\s*$"
)
CITY_STREET_LINE_RE = re.compile(r"^\s*(?P<city>[^,]+?),\s*(?P<street_part>.+?)\s*$")
POSTAL_CITY_LINE_RE = re.compile(r"^\s*(?P<postal_code>\d{2}-\d{3})\s+(?P<city>.+?)\s*$")
STREET_BUILDING_LINE_RE = re.compile(
    r"^\s*(?P<street>.+?)\s+(?P<building>\d+[A-Za-z]?(?:[-/]\d+[A-Za-z]?)?)\s*$"
)


class RawFileNotFoundError(ValueError):
    pass


class ImportBatchNotFoundError(ValueError):
    pass


class UnsupportedStagingFileTypeError(ValueError):
    pass


class InvalidRawFileContentError(ValueError):
    pass


class MissingColumnMappingError(ValueError):
    pass


class RawFileAlreadyLoadedToStagingError(ValueError):
    pass


@dataclass
class StagingLoadResult:
    import_batch_id: int
    raw_file_id: int
    entity_type: str
    records_in: int
    records_out: int
    import_status: str
    process_status: str
    missing_columns: dict[str, int]
    unrecognized_columns: dict[str, int]


def get_column_mapping(
    db: Any,
    source_system_id: int,
    entity_type: str,
    repo: Any | None = None,
) -> dict[str, str]:
    repo = repo or create_repository(db)
    return repo.get_column_mapping(source_system_id, entity_type)


def map_source_records_to_canonical(
    db: Any,
    source_system_id: int,
    entity_type: str,
    source_records: list[Mapping[str, Any]],
    repo: Any | None = None,
) -> list[dict[str, Any]]:
    mapping = get_column_mapping(db, source_system_id, entity_type, repo=repo)
    return map_records_to_canonical(source_records, mapping, entity_type)


def create_repository(db: Any) -> Any:
    from app.layers.staging_validation.repository import StagingValidationRepository

    return StagingValidationRepository(db)


def parse_raw_file_records(file_type: str, file_content: bytes | None) -> list[dict[str, Any]]:
    if not file_content:
        raise InvalidRawFileContentError("Raw file content is empty.")

    normalized_file_type = file_type.strip().upper()

    if normalized_file_type == "CSV":
        # Dekodujemy CSV przez utf-8-sig, żeby BOM nie zepsuł nazwy pierwszej kolumny
        text = file_content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise InvalidRawFileContentError("CSV file has no header row.")
        return [dict(row) for row in reader]

    if normalized_file_type == "JSON":
        # Przyjmujemy obiekt i listę JSON, żeby staging działał dla pojedynczego rekordu i paczki
        parsed = json.loads(file_content.decode("utf-8-sig"))
        if isinstance(parsed, list):
            if not all(isinstance(record, Mapping) for record in parsed):
                raise InvalidRawFileContentError("JSON array must contain objects.")
            return [dict(record) for record in parsed]
        if isinstance(parsed, Mapping):
            return [dict(parsed)]
        raise InvalidRawFileContentError("JSON root must be an object or an array.")

    if normalized_file_type == "XLSX":
        return parse_xlsx_records(file_content)

    if normalized_file_type == "XML":
        return parse_xml_records(file_content)

    raise UnsupportedStagingFileTypeError(
        f"Unsupported staging file type '{file_type}'. Supported types: CSV, JSON, XLSX, XML."
    )


def parse_xlsx_records(file_content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise UnsupportedStagingFileTypeError(
            "XLSX staging load requires openpyxl to be installed."
        ) from exc

    try:
        # Otwieramy XLSX w trybie read_only, żeby większe pliki testowe nie zużywały zbędnej pamięci
        workbook = load_workbook(
            filename=io.BytesIO(file_content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise InvalidRawFileContentError("Invalid XLSX file.") from exc

    try:
        sheet = workbook.active
        rows = [
            list(row)
            for row in sheet.iter_rows(values_only=True)
            if any(cell is not None and cell != "" for cell in row)
        ]
    finally:
        workbook.close()

    if not rows:
        raise InvalidRawFileContentError("XLSX file has no rows.")

    headers = [str(header).strip() if header is not None else "" for header in rows[0]]
    if not any(headers):
        raise InvalidRawFileContentError("XLSX header row is empty.")

    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        records.append(record)

    return records


def parse_xml_records(file_content: bytes) -> list[dict[str, Any]]:
    try:
        root = ElementTree.fromstring(file_content)
    except ElementTree.ParseError as exc:
        raise InvalidRawFileContentError("Invalid XML file.") from exc

    records: list[dict[str, Any]] = []
    for record_element in root.findall(".//record"):
        record: dict[str, Any] = {}
        fields = list(record_element.findall("field"))

        if fields:
            # Czytamy nazwy XML z atrybutu name, żeby obsłużyć kolumny ze spacjami i polskimi znakami
            for field in fields:
                field_name = field.attrib.get("name")
                if field_name:
                    record[field_name] = field.text or ""
        else:
            for child in list(record_element):
                record[child.tag] = child.text or ""

        if record:
            records.append(record)

    if not records:
        raise InvalidRawFileContentError("XML file has no record elements.")

    return records


def load_raw_file_to_staging(
    db: Any,
    raw_file_id: int,
    entity_type: str,
    repo: Any | None = None,
) -> StagingLoadResult:
    entity_type = normalize_entity_type(entity_type)
    repo = repo or create_repository(db)
    process_log = None
    batch = None

    try:
        raw_file = repo.get_raw_file(raw_file_id)
        if raw_file is None:
            raise RawFileNotFoundError(f"RawFile_ID={raw_file_id} does not exist.")

        batch = repo.get_import_batch(raw_file.ImportBatch_ID)
        if batch is None:
            raise ImportBatchNotFoundError(
                f"ImportBatch_ID={raw_file.ImportBatch_ID} does not exist."
            )

        existing_staging_records = repo.count_staging_records_for_raw_file(
            raw_file.RawFile_ID,
            entity_type,
        )
        if existing_staging_records:
            # Blokujemy drugi staging-load tego samego pliku, żeby nie zdublować rekordów w stagingu
            raise RawFileAlreadyLoadedToStagingError(
                f"RawFile_ID={raw_file.RawFile_ID} is already loaded to {entity_type} staging "
                f"({existing_staging_records} records)."
            )

        batch = repo.update_import_batch_status(batch, "PROCESSING")
        process_log = repo.create_staging_process_log(
            import_batch_id=batch.ImportBatch_ID,
            raw_file_id=raw_file.RawFile_ID,
        )

        source_records = sanitize_source_records(
            parse_raw_file_records(raw_file.File_Type, raw_file.File_Content)
        )
        mapping = repo.get_column_mapping(batch.SourceSystem_ID, entity_type)
        if not mapping:
            # Wymagamy mapowania kolumn, żeby odróżnić świadomie pominięte pola od błędu konfiguracji
            raise MissingColumnMappingError(
                f"No ColumnMapping rows for SourceSystem_ID={batch.SourceSystem_ID} and Entity_Type={entity_type}."
            )

        canonical_records = map_records_to_canonical(source_records, mapping, entity_type)
        staging_records = [
            build_staging_record(
                canonical_record=canonical_record,
                source_record=source_record,
                import_batch_id=batch.ImportBatch_ID,
                raw_file_id=raw_file.RawFile_ID,
                entity_type=entity_type,
                row_number=row_number,
            )
            for row_number, (canonical_record, source_record) in enumerate(
                zip(canonical_records, source_records),
                start=1,
            )
        ]

        if entity_type == "PERSON":
            records_out = repo.insert_person_staging_records(staging_records)
        else:
            records_out = repo.insert_party_staging_records(staging_records)

        repo.finish_process_log(
            process_log,
            status="SUCCESS",
            records_in=len(source_records),
            records_out=records_out,
        )
        batch = repo.update_import_batch_status(batch, "STAGING_LOADED", finish=True)

        return StagingLoadResult(
            import_batch_id=batch.ImportBatch_ID,
            raw_file_id=raw_file.RawFile_ID,
            entity_type=entity_type,
            records_in=len(source_records),
            records_out=records_out,
            import_status=batch.Import_Status,
            process_status="SUCCESS",
            missing_columns=count_report_columns(canonical_records, MISSING_COLUMNS_KEY),
            unrecognized_columns=count_report_columns(canonical_records, UNRECOGNIZED_COLUMNS_KEY),
        )

    except RawFileAlreadyLoadedToStagingError:
        if hasattr(repo, "rollback"):
            repo.rollback()
        # Cofamy transakcję bez FAILED, żeby powtórzone wywołanie nie psuło statusu poprawnego importu
        raise

    except Exception as exc:
        error_message = str(exc)
        if hasattr(repo, "rollback"):
            repo.rollback()
        if process_log is not None:
            repo.finish_process_log(process_log, status="FAILED", error_message=error_message)
        if batch is not None:
            repo.update_import_batch_status(
                batch,
                "FAILED",
                error_message=error_message,
                finish=True,
            )
        raise


def build_staging_record(
    canonical_record: Mapping[str, Any],
    source_record: Mapping[str, Any],
    import_batch_id: int,
    raw_file_id: int,
    entity_type: str,
    row_number: int,
) -> dict[str, Any]:
    canonical_record = sanitize_mapping(canonical_record)
    source_record = sanitize_mapping(source_record)
    canonical_columns = (
        PERSON_CANONICAL_COLUMNS
        if entity_type == "PERSON"
        else PARTY_CANONICAL_COLUMNS
    )
    staging_record = {
        column: canonical_record.get(column)
        for column in canonical_columns
        if column not in {MISSING_COLUMNS_KEY, UNRECOGNIZED_COLUMNS_KEY}
    }
    staging_record["ImportBatch_ID"] = import_batch_id
    staging_record["RawFile_ID"] = raw_file_id
    staging_record["Source_Record_ID"] = detect_source_record_id(
        canonical_record,
        source_record,
        row_number,
    )
    split_address_fields(staging_record)

    if entity_type == "PERSON":
        staging_record["Birth_Date"] = parse_date_value(staging_record.get("Birth_Date"))
    else:
        # Normalizujemy wszystkie daty PARTY jednym przebiegiem, żeby insert dostał wartości typu date
        for column in PARTY_DATE_COLUMNS:
            staging_record[column] = parse_date_value(staging_record.get(column))
        staging_record["Bank_Accounts_JSON"] = normalize_json_array_value(
            staging_record.get("Bank_Accounts_JSON")
        )
        # Składamy szerokie dane KRS do JSON, żeby później zasiliły tabele relacji i ról
        related_persons_json = extract_related_persons_json(source_record)
        related_parties_json = extract_related_parties_json(source_record)
        if related_persons_json is not None:
            staging_record["Related_Persons_JSON"] = related_persons_json
        if related_parties_json is not None:
            staging_record["Related_Parties_JSON"] = related_parties_json

    return staging_record


def sanitize_source_records(source_records: list[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return [sanitize_mapping(source_record) for source_record in source_records]


def sanitize_mapping(source_record: Mapping[str, Any]) -> dict[str, Any]:
    return {
        column_name: sanitize_value(value)
        for column_name, value in source_record.items()
    }


def sanitize_value(value: Any) -> Any:
    if isinstance(value, str):
        return sanitize_text(value)
    if isinstance(value, Mapping):
        return sanitize_mapping(value)
    if isinstance(value, list):
        return [sanitize_value(item) for item in value]
    if isinstance(value, tuple):
        return tuple(sanitize_value(item) for item in value)
    return value


def sanitize_text(value: str) -> str:
    # Czyścimy wartości przed stagingiem, żeby downstream nie porównywał zbędnych znaków
    return "".join(character for character in value if character.isprintable()).strip()


def split_address_fields(staging_record: dict[str, Any]) -> None:
    for column in ("Street", "Postal_City", "City"):
        address_line = get_clean_string(staging_record.get(column))
        if not address_line:
            continue

        if split_full_address_line(staging_record, address_line, column):
            continue

        if column == "Street":
            split_street_line(staging_record, address_line)


def split_full_address_line(
    staging_record: dict[str, Any],
    address_line: str,
    source_column: str,
) -> bool:
    for pattern in (FULL_ADDRESS_STREET_FIRST_RE, FULL_ADDRESS_POSTAL_FIRST_RE):
        full_match = pattern.match(address_line)
        if full_match:
            # Rozbijamy pełną linię adresu, żeby staging nie trzymał miasta i kodu w ulicy
            set_address_part(
                staging_record,
                "Postal_Code",
                full_match.group("postal_code").strip(),
                source_column,
            )
            set_address_part(
                staging_record,
                "City",
                full_match.group("city").strip(),
                source_column,
            )
            split_street_line(staging_record, full_match.group("street_part"))
            clear_embedded_address_source(staging_record, source_column)
            return True

    city_street_match = CITY_STREET_LINE_RE.match(address_line)
    if city_street_match and looks_like_street_line(city_street_match.group("street_part")):
        # Obsługujemy przypadek miasto-ulica, gdy kod pocztowy przyszedł już osobną kolumną
        set_address_part(
            staging_record,
            "City",
            city_street_match.group("city").strip(),
            source_column,
        )
        split_street_line(staging_record, city_street_match.group("street_part"))
        clear_embedded_address_source(staging_record, source_column)
        return True

    postal_city_match = POSTAL_CITY_LINE_RE.match(address_line)
    if postal_city_match:
        # Rozdzielamy linię kod-miasto, żeby adres bez ulicy nie udawał nazwy ulicy
        set_address_part(
            staging_record,
            "Postal_Code",
            postal_city_match.group("postal_code"),
            source_column,
        )
        set_address_part(
            staging_record,
            "City",
            postal_city_match.group("city"),
            source_column,
        )
        if source_column == "Street":
            staging_record["Street"] = None
        return True

    return False


def split_street_line(staging_record: dict[str, Any], street_line: str) -> None:
    normalized_line = normalize_street_line(street_line)
    if not normalized_line:
        staging_record["Street"] = None
        return

    street_match = STREET_BUILDING_LINE_RE.match(normalized_line)
    if not street_match:
        staging_record["Street"] = normalized_line
        return

    # Oddzielamy nazwę ulicy od numeru, żeby finalny DIM_ADDRESS dostał osobne pola adresowe
    building_number, apartment_number = split_building_and_apartment(
        street_match.group("building")
    )
    staging_record["Street"] = street_match.group("street").strip()
    staging_record["Building_Number"] = staging_record.get("Building_Number") or building_number
    staging_record["Apartment_Number"] = staging_record.get("Apartment_Number") or apartment_number


def normalize_street_line(value: str) -> str:
    without_prefix = ADDRESS_PREFIX_RE.sub("", value).strip()
    return APARTMENT_PREFIX_RE.sub("/", without_prefix).strip()


def looks_like_street_line(value: str) -> bool:
    return STREET_BUILDING_LINE_RE.match(normalize_street_line(value)) is not None


def clear_embedded_address_source(staging_record: dict[str, Any], source_column: str) -> None:
    if source_column == "Street":
        return
    if source_column == "City":
        staging_record[source_column] = staging_record.get("City")
    else:
        staging_record[source_column] = None


def set_address_part(
    staging_record: dict[str, Any],
    target_column: str,
    value: str,
    source_column: str,
) -> None:
    current_value = staging_record.get(target_column)
    if source_column == target_column or current_value in (None, ""):
        staging_record[target_column] = value


def get_clean_string(value: Any) -> str | None:
    if value in (None, ""):
        return None
    cleaned_value = str(value).strip()
    return cleaned_value or None


def split_building_and_apartment(value: str) -> tuple[str, str | None]:
    if "/" not in value:
        return value, None

    # Slash w danych testowych oznacza numer lokalu, więc zapisujemy go osobno od posesji
    building_number, apartment_number = value.split("/", 1)
    return building_number.strip(), apartment_number.strip() or None


def detect_source_record_id(
    canonical_record: Mapping[str, Any],
    source_record: Mapping[str, Any],
    row_number: int,
) -> str:
    canonical_source_record_id = canonical_record.get("Source_Record_ID")
    if canonical_source_record_id not in (None, ""):
        return str(canonical_source_record_id)

    # Wybieramy stabilne ID rekordu, żeby lineage nie opierał się wyłącznie na numerze wiersza
    lookup = {
        str(column_name).strip().casefold(): value
        for column_name, value in source_record.items()
    }
    for candidate in SOURCE_RECORD_ID_CANDIDATES:
        value = lookup.get(candidate.strip().casefold())
        if value not in (None, ""):
            return str(value)

    return str(row_number)


def parse_date_value(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    try:
        # Bierzemy pierwsze 10 znaków daty, żeby ISO datetime zapisać jako czysty DATE
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def normalize_json_array_value(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        try:
            parsed = json.loads(stripped)
        except json.JSONDecodeError:
            # Zamieniamy tekst kont VAT na listę, żeby staging zawsze trzymał poprawny JSON array
            parsed = [part.strip() for part in stripped.split(",") if part.strip()]
    else:
        parsed = value

    if isinstance(parsed, list):
        return json.dumps(parsed, ensure_ascii=False, default=str)
    return json.dumps([parsed], ensure_ascii=False, default=str)


def extract_related_persons_json(source_record: Mapping[str, Any]) -> str | None:
    # Zbieramy osoby z KRS do listy, żeby goldenizacja mogła budować role person-party
    persons_by_key: dict[tuple[str, int], dict[str, Any]] = {}
    field_map = {
        "Imie": "first_name",
        "Nazwisko": "last_name",
        "PESEL": "pesel",
        "Funkcja": "role_name",
        "DataOd": "valid_from",
        "DataDo": "valid_to",
    }

    for source_column, value in source_record.items():
        if value in (None, ""):
            continue
        match = KRS_PERSON_ROLE_COLUMN_RE.match(str(source_column))
        if not match:
            continue

        prefix, slot, field_name = match.groups()
        key = (prefix, int(slot))
        person = persons_by_key.setdefault(
            key,
            {
                "role_group": KRS_PERSON_ROLE_PREFIXES[prefix],
                "slot": int(slot),
            },
        )
        person[field_map[field_name]] = value

    return compact_json_or_none(persons_by_key.values())


def extract_related_parties_json(source_record: Mapping[str, Any]) -> str | None:
    # Zbieramy podmioty z KRS osobno, żeby goldenizacja mogła budować relacje party-party
    parties_by_key: dict[tuple[str, int], dict[str, Any]] = {}
    field_map = {
        "Nazwa": "name",
        "KRS": "krs",
        "NIP": "nip",
        "DataOd": "valid_from",
        "DataDo": "valid_to",
    }

    for source_column, value in source_record.items():
        if value in (None, ""):
            continue
        match = KRS_PARTY_RELATIONSHIP_COLUMN_RE.match(str(source_column))
        if not match:
            continue

        prefix, slot, field_name = match.groups()
        key = (prefix, int(slot))
        party = parties_by_key.setdefault(
            key,
            {
                "relationship_group": "PARTY_SHAREHOLDER",
                "slot": int(slot),
            },
        )
        party[field_map[field_name]] = value

    return compact_json_or_none(parties_by_key.values())


def compact_json_or_none(items: Iterable[Mapping[str, Any]]) -> str | None:
    # Usuwamy puste sloty KRS, żeby JSON zawierał tylko relacje z realnymi danymi
    compact_items = [
        {
            key: value
            for key, value in item.items()
            if value not in (None, "")
        }
        for item in items
    ]
    compact_items = [
        item
        for item in compact_items
        if any(key != "slot" and value not in (None, "") for key, value in item.items())
    ]
    if not compact_items:
        return None
    compact_items.sort(
        key=lambda item: (
            str(item.get("role_group") or item.get("relationship_group")),
            item["slot"],
        )
    )
    return json.dumps(compact_items, ensure_ascii=False, default=str)


def count_report_columns(
    canonical_records: list[Mapping[str, Any]],
    report_key: str,
) -> dict[str, int]:
    counter: Counter[str] = Counter()
    for record in canonical_records:
        for column in record.get(report_key, []):
            column_name = str(column)
            if (
                report_key == UNRECOGNIZED_COLUMNS_KEY
                and column_name.strip().casefold() in SOURCE_RECORD_ID_CANDIDATE_KEYS
            ):
                # Pomijamy ID źródłowe w raporcie, żeby unrecognized pokazywało realne braki mapowania
                continue
            counter.update([column_name])
    return dict(counter)
